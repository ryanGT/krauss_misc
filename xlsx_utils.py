import openpyxl, datetime, os, txt_mixin
import pandas as pd


def search_one_label(cellin, match):
    value = cellin.value
    if type(value) == unicode:
        labelin = value.encode()
    elif type(value) == str:
        labelin = value
    else:
        #not sure what to do with other data types
        return False

    search_label = labelin.lower()
    search_label = search_label.replace(' ','')
    return search_label == match


def find_label_row(sheet):
    for i, row in enumerate(sheet.iter_rows()):
        bool1 = search_one_label(row[1], 'lastname')
        bool2 = search_one_label(row[2], 'firstname')
        if bool1 and bool2:
            return row, i


def get_row_values(rowin):
    values = [item.value for item in rowin]
    return values


def encode_values(list_of_values):
    values_out = []

    for item in list_of_values:
        if type(item) == str:
            value_out = item.encode('ascii', 'ignore')
        elif type(item) == datetime.datetime:
            fmt_time = '%m/%d/%y'#<-- this seems like a risky hard code
            value_out = item.strftime(fmt_time)
        elif item is None:
            value_out = ''
        else:
            value_out = item
        values_out.append(value_out)

    return values_out


def convert_row_to_string(rowin):
    values = get_row_values(rowin)
    encoded_values = encode_values(values)
    return encoded_values


def clean_string(str_in):
    """clean up strings that I find hard to deal with in spreadsheet
    cells (mainly newlines)."""
    if type(str_in) in [int, float, bool]:
        return str_in
    else:
        #print('type(str_in) = %s' % type(str_in))
        #str_in = str(str_in)
        if type(str_in) == bytes:
            str_in = str_in.decode("utf-8")
    find_rep_tuples = [('\n','; '), \
                       ('\r','; '), \
                       ]
    str_out = str_in

    for find_str, rep_str in find_rep_tuples:
        str_out = str_out.replace(find_str, rep_str)

    return str_out


def clean_encoded_strings(list_in):
    list_out = [clean_string(item) for item in list_in]
    return list_out


def quote_strings(listin):
    listout = []
    for item in listin:
        if type(item) in [str, unicode]:
            value_out = '"%s"' % item
        else:
            value_out = item
        listout.append(value_out)
    return listout


def empty_row_check(row_in, N=3):
    for i in range(N):
        if row_in[i].internal_value is not None:
            return False

    return True


def get_all_data(sheet_in, debug=0):
    """One trick is that empty but formatted rows show up in the rows
    of the sheet.  I will break my for loop on the first empty row
    after the label rows."""
    ## label_row, label_ind = find_label_row(sheet_in)
    ## encoded_labels = convert_row_to_string(label_row)
    #N = len(sheet_in.rows)
    ## start_row = label_ind + 1
    start_row = 0
    
    data_out = []
    
    for cur_row in sheet_in.rows:
        #cur_row = sheet_in.rows[i]
        #if debug > 0:
        #    print('i = %i' % i)
        if empty_row_check(cur_row):
            break
        
        cur_values = get_row_values(cur_row)
        if debug > 0:
            print('cur_values = ' + str(cur_values))
        encoded_values = encode_values(cur_values)
        if debug > 0:
            print('encoded_values = ' + str(encoded_values))

        clean_values = clean_encoded_strings(encoded_values)
        data_out.append(clean_values)

    return data_out



def get_all_data_by_filename(filename, sheet_ind=0, verbosity=1, \
                             data_only=False):
    wb = openpyxl.load_workbook(filename, data_only=data_only)
    sheet_names = wb.sheetnames

    if len(sheet_names) > 1 and verbosity > 0:
        print("Warning: more than one sheet")

    sheet1 = wb[sheet_names[0]]

    data_out = get_all_data(sheet_in=sheet1)

    return data_out


def xlsx_to_csv(xlsx_name):
    pne, ext = os.path.splitext(xlsx_name)
    csv_path = pne + '.csv'
    df = pd.DataFrame(pd.read_excel(xlsx_name))
    df.to_csv(csv_path, index=False)
    #data = get_all_data_by_filename(xlsx_name)
    #txt_mixin.dump_delimited(csv_path, data, delim='\t')
    return csv_path
